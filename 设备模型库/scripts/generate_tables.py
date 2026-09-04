# -*- coding: utf-8 -*-
"""表生成器: 从 SQLite 设备模型库生成 变量表CSV + 通讯字表XLSX + 结构体/FB表."""
import os, sqlite3, sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from shared_data import STRUCTURES, FBS, CON_VARS, ARCHETYPE_EXTRA, HMI_BASE, SHARED_CONST

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE, 'db', 'model.db')
OUT = os.path.join(BASE, 'output')

# GBK 表头 (汇川规范)
HEADER_VAR = '序号,变量名,数据类型,隐藏初始值,初始值,掉电保持,网络公开,注释,软元件地址,'
HEADER_STRUCT = '序号,成员变量名,数据类型,注释,'
HEADER_FB = '序号,类别,名称,数据类型,隐藏初始值,初始值,掉电保持,注释,'
HEADER_INST = '序号,变量名,数据类型,隐藏初始值,初始值,注释'


def gbk_crlf(text):
    return text.replace('\r\n', '\n').replace('\n', '\r\n').encode('gbk')


def write_csv(path, header, rows):
    lines = [header] + rows
    data = '\r\n'.join(lines)
    with open(path, 'wb') as f:
        f.write(gbk_crlf(data))
    print('[gen] %s (%d 行)' % (os.path.basename(path), len(rows)))


def get_archetype(con, project):
    return con.execute('SELECT archetype FROM project WHERE name=?', (project,)).fetchone()[0]


def gen_variable_tables(con, out_dir, project):
    """生成 io / host 变量表 (host 追加原型扩展)."""
    archetype = get_archetype(con, project)
    pid = con.execute('SELECT id FROM project WHERE name=?', (project,)).fetchone()[0]
    # io 表
    rows = []
    n = 1
    for v in con.execute('SELECT var_name, data_type, address, comment, filter_mode FROM variable '
                         'WHERE category="io" AND project_id=? ORDER BY rowid', (pid,)):
        name, dt, addr, comment, fm = v
        if comment and fm and fm != 'builtin':
            comment += '(滤波:%s)' % fm
        rows.append('%d,%s,%s,,OFF,,内部,%s,%s,' % (n, name, dt, comment or '', addr or ''))
        n += 1
    write_csv(os.path.join(out_dir, 'VariableTable_io.csv'), HEADER_VAR, rows)

    # host 表 (本工程通讯寄存器 -> host_Send_*/host_Rcv_*)
    regs = list(con.execute('SELECT var_name, data_type, d_addr, name_cn, direction FROM comm_reg '
                            'WHERE project_id=? ORDER BY d_addr', (pid,)))
    rows = []
    n = 1
    for var, dt, d, cn, direction in regs:
        kind = '命令' if direction == 'write' else '状态'
        rows.append('%d,%s,%s,0,0,,公开,%s,D%d,' % (n, var, dt, cn, d))
        n += 1
    # 追加原型扩展 host 变量 (D 地址顺延)
    next_d = max((r[2] + (2 if r[1].upper() in ('REAL', 'DINT') else 1) for r in regs), default=0)
    for var, dt, cn, direction in ARCHETYPE_EXTRA.get(archetype, {}).get('host', []):
        kind = '命令' if direction == 'write' else '状态'
        rows.append('%d,%s,%s,0,0,,公开,%s,D%d,' % (n, var, dt, cn, next_d))
        n += 1
        next_d += 2 if dt in ('REAL', 'DINT') else 1
    write_csv(os.path.join(out_dir, 'VariableTable_host.csv'), HEADER_VAR, rows)


def gen_hmi_table(con, out_dir, project):
    """HMI 手动变量表 (按设备推导, 通用)."""
    pid = con.execute('SELECT id FROM project WHERE name=?', (project,)).fetchone()[0]
    rows = []
    n = 1
    s = HMI_BASE
    for r in con.execute('SELECT idx FROM actuator WHERE project_id=? AND kind="servo" ORDER BY idx', (pid,)):
        idx = r[0]
        rows.append('%d,hmi_JogP_M%d,BOOL,,OFF,,内部,%d轴正向点动,S%d,' % (n, idx, idx, s)); n += 1; s += 1
        rows.append('%d,hmi_JogN_M%d,BOOL,,OFF,,内部,%d轴反向点动,S%d,' % (n, idx, idx, s)); n += 1; s += 1
    for r in con.execute('SELECT eng_name FROM actuator WHERE project_id=? AND kind="stepper" ORDER BY idx', (pid,)):
        eng = r[0]
        rows.append('%d,hmi_StepP_%s,BOOL,,OFF,,内部,%s正向点动,S%d,' % (n, eng, eng, s)); n += 1; s += 1
        rows.append('%d,hmi_StepN_%s,BOOL,,OFF,,内部,%s反向点动,S%d,' % (n, eng, eng, s)); n += 1; s += 1
    if con.execute('SELECT 1 FROM actuator WHERE project_id=? AND kind="roller"', (pid,)).fetchone():
        rows.append('%d,hmi_Roller_Fwd,BOOL,,OFF,,内部,辊筒正转,S%d,' % (n, s)); n += 1; s += 1
        rows.append('%d,hmi_Roller_Rev,BOOL,,OFF,,内部,辊筒反转,S%d,' % (n, s)); n += 1; s += 1
    write_csv(os.path.join(out_dir, 'VariableTable_hmi.csv'), HEADER_VAR, rows)


def gen_comm_xlsx(con, out_dir):
    """生成通讯字表 XLSX (美化, 同示例样式)."""
    # 样式
    header_font = Font(name='华文楷体', size=16, bold=True)
    header_fill = PatternFill(fill_type='solid', fgColor='FF00B0F0')
    label_font = Font(name='楷体', size=18, bold=True)
    read_fill = PatternFill(fill_type='solid', fgColor='FF92D050')
    write_fill = PatternFill(fill_type='solid', fgColor='FFFFFF00')
    data_font = Font(name='等线', size=11)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    label_align = Alignment(horizontal='center', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '通讯字表'
    for col, w in zip('ABCDE', [16, 35.33, 11, 9.25, 66.25]):
        ws.column_dimensions[col].width = w

    headers = ['服务器读写', '通讯字名称', 'Mobus地址', '服务器地址', '功能示意']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = header_align
    ws.row_dimensions[1].height = 45

    def section(start_row, items, label, fill):
        for i, (name, addr, off, desc) in enumerate(items):
            r = start_row + i
            for col, val, al in ((2, name, left_align), (3, addr, center_align),
                                 (4, off, center_align), (5, desc, left_align)):
                c = ws.cell(row=r, column=col, value=val)
                c.font = data_font; c.border = border; c.alignment = al
        end = start_row + len(items) - 1
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end, end_column=1)
        lab = ws.cell(row=start_row, column=1, value=label)
        lab.font = label_font; lab.fill = fill; lab.alignment = label_align
        for r in range(start_row, end + 1):
            ws.cell(row=r, column=1).border = border
        return end

    # 通讯寄存器 -> (name_cn, modbus, d_addr+1, comment)
    write_items, read_items = [], []
    for r in con.execute('SELECT direction, name_cn, data_type, d_addr, modbus_addr, comment '
                         'FROM comm_reg ORDER BY d_addr'):
        direction, cn, dt, d, mb, comment = r
        # 服务器地址 = 起始偏移 (相对 40001); Modbus 已存
        off = d
        item = (cn, mb, off, comment or '')
        (read_items if direction == 'read' else write_items).append(item)

    # 地址升序: 写段(D0起)在前, 读段(D17起)在后
    row = 2
    if write_items:
        row = section(row, write_items, '服务器写', write_fill) + 1
    if read_items:
        section(row, read_items, '服务器读', read_fill)

    path = os.path.join(out_dir, '通讯字表.xlsx')
    wb.save(path)
    print('[gen] %s (写=%d 读=%d)' % (os.path.basename(path), len(write_items), len(read_items)))


def gen_structure_tables(out_dir):
    """结构体表 (共享定义)."""
    for name, members in STRUCTURES.items():
        rows = ['%d,%s,%s,%s,' % (i, m, t, c) for i, (m, t, c) in enumerate(members, 1)]
        write_csv(os.path.join(out_dir, 'StructureTable_%s.csv' % name), HEADER_STRUCT, rows)
    print('[gen] 结构体表 x%d' % len(STRUCTURES))


def gen_fb_tables(out_dir):
    """FB 表 (共享定义)."""
    for name, params in FBS.items():
        rows = ['%d,%s,%s,%s,,FALSE,,%s,' % (i, cat, p, t, c) for i, (cat, p, t, c) in enumerate(params, 1)]
        write_csv(os.path.join(out_dir, 'FbTable_%s.csv' % name), HEADER_FB, rows)
    print('[gen] FB 表 x%d' % len(FBS))


def gen_con_table(con, out_dir, project):
    """con 内部控制变量表 + 轴实体/句柄 + 原型扩展."""
    archetype = get_archetype(con, project)
    pid = con.execute('SELECT id FROM project WHERE name=?', (project,)).fetchone()[0]
    rows = []
    n = 1
    for name, dt, addr, comment in CON_VARS:
        rows.append('%d,%s,%s,,OFF,,内部,%s,%s,' % (n, name, dt, comment, addr))
        n += 1
    # 原型扩展 con 变量 (D 地址顺延)
    next_d = int(CON_VARS[-1][2][1:]) + 1
    for name, dt, comment in ARCHETYPE_EXTRA.get(archetype, {}).get('con', []):
        rows.append('%d,%s,%s,,OFF,,内部,%s,D%d,' % (n, name, dt, comment, next_d))
        n += 1
        next_d += 1
    for r in con.execute('SELECT cmd_struct, status_struct, fb_instance, axis_handle, idx, eng_name, kind FROM actuator '
                         'WHERE project_id=? AND kind IN ("servo","stepper") ORDER BY kind, idx', (pid,)):
        cmd, status, fb, ah, idx, eng, kind = r
        rows.append('%d,%s,%s,,,内部,%sM%d%s命令结构实体,,' % (n, cmd, 'AxisControlData', kind, idx, eng)); n += 1
        if status:
            rows.append('%d,%s,%s,,,内部,%sM%d%s状态结构实体,,' % (n, status, 'AxisStatusData', kind, idx, eng)); n += 1
    write_csv(os.path.join(out_dir, 'VariableTable_con.csv'), HEADER_VAR, rows)


def gen_inst_table(con, out_dir, project):
    """功能块实例表 (inst) + 原型扩展."""
    archetype = get_archetype(con, project)
    pid = con.execute('SELECT id FROM project WHERE name=?', (project,)).fetchone()[0]
    rows = []
    n = 1
    # 报警锁存实例 (SBR_03 故障锁存闭环, 全原型必配, 细则2)
    rows.append('%d,Latch_Timeout,FB_AlarmLatch,,,,' % n); n += 1
    # 通讯恢复沿 (SBR_host 段3 清防重放序号记忆, 细则15)
    rows.append('%d,r_CommLostF,TRIG.F_TRIG,,,,' % n); n += 1
    rows.append(',r_CommLostF.CLK,BOOL,,,,')
    rows.append(',r_CommLostF.Q,BOOL,,,,')
    rows.append(',r_CommLostF.M,BOOL,,,,')
    for r in con.execute('SELECT fb_instance, eng_name FROM actuator WHERE project_id=? AND kind="servo" ORDER BY idx', (pid,)):
        fb, eng = r
        rows.append('%d,%s,FB_EtherCAT_Axis_ST,,,,' % (n, fb)); n += 1
        rows.append(',%s.Axis,_sMCAXIS_INFO,,,,' % fb)
        # host_driven: 每伺服命令握手实例 (SBR_06, 细则3)
        if archetype == 'host_driven':
            rows.append('%d,Handshake_%s,FB_CmdHandshake,,,,' % (n, eng)); n += 1
    for r in con.execute('SELECT fb_instance FROM actuator WHERE project_id=? AND kind="stepper" ORDER BY idx', (pid,)):
        fb = r[0]
        rows.append('%d,%s,FB_StepperDrive,,,,' % (n, fb)); n += 1
    # 原型扩展 inst (R_TRIG / TON 等)
    for name, ptype, comment in ARCHETYPE_EXTRA.get(archetype, {}).get('inst', []):
        rows.append('%d,%s,%s,,,,' % (n, name, ptype)); n += 1
    write_csv(os.path.join(out_dir, '功能块实例.csv'), HEADER_INST, rows)


def gen_const_table(con, out_dir, project):
    """常量表 = 共享常量 + 规格单常量."""
    pid = con.execute('SELECT id FROM project WHERE name=?', (project,)).fetchone()[0]
    rows = []
    n = 1
    for name, dt, value, comment in SHARED_CONST:
        rows.append('%d,%s,%s,,%s,保持,,%s,,' % (n, name, dt, value, comment)); n += 1
    for r in con.execute('SELECT name, data_type, value, comment FROM const_item WHERE project_id=? ORDER BY rowid', (pid,)):
        rows.append('%d,%s,%s,,%s,保持,,%s,,' % (n, r[0], r[1], r[2], r[3] or '')); n += 1
    write_csv(os.path.join(out_dir, 'VariableTable_const.csv'), HEADER_VAR, rows)


def main(project=None):
    con = sqlite3.connect(DB_PATH)
    if project:
        con.execute('SELECT 1 FROM project WHERE name=?', (project,)).fetchone() \
            or sys.exit('工程不存在: ' + project)
    else:
        project = con.execute('SELECT name FROM project LIMIT 1').fetchone()[0]
    out_dir = os.path.join(OUT, project)
    os.makedirs(out_dir, exist_ok=True)
    gen_variable_tables(con, out_dir, project)
    gen_hmi_table(con, out_dir, project)
    gen_comm_xlsx(con, out_dir)
    gen_structure_tables(out_dir)
    gen_fb_tables(out_dir)
    gen_con_table(con, out_dir, project)
    gen_inst_table(con, out_dir, project)
    gen_const_table(con, out_dir, project)
    con.close()


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else None)
